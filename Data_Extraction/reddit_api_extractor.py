import praw
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv, find_dotenv
import os

# Load environment variables
_ = load_dotenv(find_dotenv())
client_secret = os.environ.get('reddit_client_secret')

# Authenticate to Reddit
reddit = praw.Reddit(
    client_id='uvlxFUqckCx7WpdlEalNYA',
    client_secret=client_secret,
    username='Known_Way_2607',
    password='Pq$s9-TWDD7&4,5',
    user_agent='sentimenti by u/Known_Way_2607'
)

# Define keywords related to immigration sentiment
# keywords = [
#     'accommodation', 'cultural experiences', 'work-life balance', 
#     'immigration', 'immigration issues', 'moving abroad', 'living in a new country', 
#     'expat', 'immigrant', 'housing', 'weather', 'cantonese', 
#     'racism', 'refugees', 'filipino', 'domestic helpers', 'mandarin', 
#     'culture', 'HK immigration', 'Hong Kong lifestyle', 'HK visa', 
#     'job opportunities', 'employment in Hong Kong', 'career', 
#     'local cuisine', 'entertainment', 'recreation', 
#     'visa process', 'immigration laws', 
#     'cultural challenges', 'language barriers', 
#     'expat community', 'local community',
#     'community integration', 'social life', 'public services', 
#     'education system', 'healthcare system', 'transportation', 
#     'cost of living', 'local customs', 'diversity', 
#     'quality of life', 'social issues', 'government policies', 
#     'economic opportunities', 'citizenship', 'legal status', 
#     'family life', 'relationship dynamics', 'networking', 
#     'public perception', 'integration challenges',
#     'Indian', 'Black', 'White', 'race relations', 'ethnic diversity',
#     'healthcare', 'education', 'expensive', 'jobs', 
#     'permanent residence', 'expats', 'immigrants', 'English', 
#     'family', 'politics'
# ]

keywords = ['cultural experiences',
    'immigration', 'living in a new country', 
    'expat', 'immigrant', 'housing']

# keywords = ['cultural experiences',
# 'cantonese', 'domestic helpers', 'mandarin', 
#     'culture', 'HK immigration', 'Hong Kong lifestyle', 'employment in Hong Kong', 
#   'locals', 'blacks', 'whites, social life'
# ]

# keywords = [
#     'education', 'transportation', 
#     'cost of living', 'social issues', 
#     'Indians', 'black people', 'white people', 'expensive', 'jobs', 'immigrants', 'English'
# ]


# List of subreddits related to Hong Kong
hongkong_subreddits = ['HongKong', 'HongKongProtest', 'HongKongTravel', 'HongKongExpats', 'hongkong_forhire', 'hongkongsar']

file_path = f"/Users/ngoni/Desktop/Immigration Sentiment Analysis/Data_Extraction/reddit_data4.xlsx"

# Load the existing workbook or create a new one
try:
    wb = load_workbook(file_path)
    sheet = wb.active
except FileNotFoundError:
    wb = Workbook()
    sheet = wb.active
    # Add headers if the file is newly created
    sheet.append(['Subreddit', 'Keyword', 'Comment'])


# Iterate through each keyword and subreddit
for keyword in keywords:
    for subreddit_name in hongkong_subreddits:
        subreddit = reddit.subreddit(subreddit_name)
        
        # Perform a search for the keyword in the current subreddit
        keyword_results = subreddit.search(keyword, sort='new', time_filter='year', limit=None)
        
        # Append the comments for each result to the Excel sheet
        for submission in keyword_results:
            submission.comments.replace_more(limit=None)  # Retrieve all comments, even nested ones
            
            for comment in submission.comments.list():
                # Append data to the Excel sheet
                sheet.append([subreddit_name, keyword, comment.body])

# Save the workbook
wb.save(filename=file_path)
