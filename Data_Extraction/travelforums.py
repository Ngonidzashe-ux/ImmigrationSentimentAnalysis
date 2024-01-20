import os
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
import pandas as pd

# URL of the webpage
url = 'http://www.theothersideforever.com/moving-to-a-new-country-alone/'

# Send a GET request to the URL
response = requests.get(url)

# Check if the request was successful (status code 200)
if response.status_code == 200:
    # Parse the HTML content of the page
    soup = BeautifulSoup(response.content, 'html.parser')

    # Extract blog article content
    article_content = soup.find('div', class_='entry-content')
    if article_content:
        article_content = article_content.text.strip()
    else:
        article_content = "No article content found."

    # Extract comments
    comments = soup.find_all('div', class_='commententry')
    comments_content = [comment.p.text.strip() for comment in comments]

    # Combine blog article and comments into a single list
    combined_content = [article_content] + comments_content

    # Specify the file path
    file_path = "/Users/ngoni/Desktop/Immigration Sentiment Analysis/Data_Extraction/travel_forums.xlsx"

    # Check if the workbook exists
    if os.path.exists(file_path):
        # Load the existing workbook
        wb = load_workbook(file_path)
        sheet = wb.active
    else:
        # Create a new workbook if it doesn't exist
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Content"])  # Add header row

    # Append combined content to the Excel sheet
    for content in combined_content:
        sheet.append([content])

    # Save the updated workbook
    wb.save(file_path)

    print("Data appended to Excel sheet.")
else:
    print(f"Failed to fetch the page. Status code: {response.status_code}")
